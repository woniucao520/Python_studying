# -*- coding: utf-8 -*-
import hashlib
import json
import datetime
import os
import urllib2
import requests
from lxml import etree
#import xlwt
from openpyxl import Workbook
import tempfile
import zipfile

import sys
reload(sys)
sys.setdefaultencoding("utf-8")


class Kebilin(object):

    def __init__(self, serial):
        res = self.get_product_info(serial)
        folder_path= self.create_main_folder()
        self.create_excel_file(res, folder_path)
        self.download_image(res, folder_path)
        self.down_zip(folder_path)

    def get_product_info(self, serial, page_size=100):
        appid = '25'
        secert = '4122e029f305f66dfb62849fe97f83abc7c46be1'
        timestamp = datetime.datetime.strftime(datetime.datetime.now(), '%Y-%m-%d %H:%M:%S')  # '2017-08-16 11:11:12'

        sign = appid + secert + timestamp

        md5_encode = hashlib.md5()
        md5_encode.update(sign)
        sign = md5_encode.hexdigest()

        data = {
            'head': {
                'userid': appid,
                'timestamp': timestamp,
                'sign': sign,
                'f': 'get_goods_list'
            },
            'body': {
                'serial': serial,
                'page_no': '',
                'page_size': page_size,
                'starttime': '',
                'endtime': '',
            }
        }

        url = 'http://api.kebilin.com'
        data = json.dumps(data)
        req = urllib2.Request(url, data, {'Content-Type': 'application/json'})
        response = urllib2.urlopen(req)
        res = response.read() # 获取网页源码
        res = json.loads(res)  # json.loads:str转成dict

        return res

    def create_main_folder(self):
        # 创建 文件夹的时间
        mkfile_time = datetime.datetime.strftime(datetime.datetime.now(), '%Y%m%d%H%M%S')
        mkfile_time = 'kbl'+mkfile_time

        # if os.name == 'posix':
        #     mypath = './' + mkfile_time
        # else:
        #     mypath='D:\\' + mkfile_time  # 指定放在D 盘 目录下
        # if not os.path.exists(mypath):
        #     os.makedirs(mypath)  # makedirs 创建多级目录文件夹，mkdir创建一个文件夹
        #
        # return mypath
        # 创建临时文件夹
        t = tempfile.mkdtemp()   # 创建临时目录
        # print (mypath)   #  C:\Users\wineplus\AppData\Local\Temp\tmprbdtogsa
        mypath = t + os.sep
        mypath = mypath + mkfile_time
        if not os.path.exists(mypath):
            os.makedirs(mypath)
            os.system("explorer.exe " + t)
        return mypath



    def create_excel_file(self, res, mypath):


        result = res['root']['body']
        # 实例化一个Workbook()对象(即excel文件)
        # wbk = xlwt.Workbook()
        wbk = Workbook()
        ws = wbk.active
        # 新建一个名为Sheet1的excel sheet。此处的cell_overwrite_ok =True是为了能对同一个单元格重复操作。
        #sheet = wbk.add_sheet('Sheet1',cell_overwrite_ok=True)
        # 遍历result中的没个元素。
        header_list = ['vendor_code', 'gallery', 'description', 'weight', 'type_id', 'color', 'image',
                       'product_sync_type', 'price', 'special_price', 'qty', 'short_description', 'size',
                       'sku', 'name', 'brand', 'categories', 'attribute_set', 'country', 'warehouse', 'barcode']
        ws.append(header_list)
        #index = 0
        #for data in header_list:
            #sheet.write(0, index, data)
            #index += 1

        #row = 1
        for r in result:
            images = 'images' in r.keys() and r['images'] or ''
            content = 'content' in r.keys() and r['content'] or ''
            price = 'price' in r.keys() and r['price'] or 0
            name = 'goods_name' in r.keys() and r['goods_name'] or ''
            barcode = 'barcode' in r.keys() and r['barcode'] or ''
            country = 'madein' in r.keys() and r['madein'] or ''
            if country == u'台湾' or country == u'香港':
                country = u'中国' + country

            sku = 'sn' in r.keys() and 'kbl' + r['sn'].lower() or ''
            # 获取重量
            skus = 'skus' in r.keys() and  r['skus'] or ''

            if skus and ('sku' in skus.keys()) and skus['sku']:
                weight = skus['sku']['weight'] or 0.5
                qty = skus['sku']['stock'] or 0
                qty = qty > 10 and qty - 10 or 0
            else:
                weight = 0.5
                qty = 0

            vendor_code = 9000
            type_id = 'simple'
            product_sync_type = u'售卖且不需要同步'
            warehouse = u'青岛1号仓'
            default_image = sku + '_8001.jpg'

            gallery = ''
            description = ''

            for image in range(0, len(images)):
                if gallery:
                    gallery += '||'

                gallery += sku + '_800' + str(image + 1) + '.jpg'

            content_list = content.split('<img')
            for con in range(1, len(content_list)-1):
                content_add = '<img src="{{media url="wysiwyg/product/k/b/' + sku + '/' + sku + '_' + str(con).zfill(2) + '.jpg"}}" alt="' + name + '" />'
                description += content_add

            """
            sheet.write(row, 0, vendor_code)
            sheet.write(row, 1, gallery)
            sheet.write(row, 2, description)
            sheet.write(row, 3, weight)
            sheet.write(row, 4, type_id)
            sheet.write(row, 5, '')
            sheet.write(row, 6, default_image)
            sheet.write(row, 7, product_sync_type)
            sheet.write(row, 8, price)
            sheet.write(row, 9, '')
            sheet.write(row, 10, qty)
            sheet.write(row, 11, '')
            sheet.write(row, 12, '')
            sheet.write(row, 13, sku)
            sheet.write(row, 14, name)
            sheet.write(row, 15, '')
            sheet.write(row, 16, '')
            sheet.write(row, 17, '')
            sheet.write(row, 18, country)
            sheet.write(row, 19, warehouse)
            sheet.write(row, 20, barcode)
            """

            rows = []
            rows.append(vendor_code)

            rows.append(gallery)
            rows.append(description)
            rows.append(weight)
            rows.append(type_id)
            rows.append('')

            rows.append(default_image)
            rows.append(product_sync_type)
            rows.append(price)
            rows.append('')
            rows.append(qty)

            rows.append('')
            rows.append('')
            rows.append(sku)
            rows.append(name)
            rows.append('')

            rows.append('')
            rows.append('')
            rows.append(country)
            rows.append(warehouse)
            rows.append(barcode)

            ws.append(rows)
            #row += 1

        # 以传递的name+当前日期作为excel名称保存。
        wbk.save(mypath + '/kebilin_products.xlsx')

    def download_image(self, res, mkfile_time):
        for d in res['root']['body']:
            description= d['content']
            su = d['sn']   # 获取sku

            new_sku = su.lower()  # 直接一步转换小写到位
            new_sku = 'kbl'+ new_sku

            des_tree = etree.HTML(description)
            des_list = des_tree.xpath('//img/@src')

            i = 1
            for list in des_list:
                list = ''.join(list) # 转换成str
                if 'top.jpg' in list:  # 判断字符串里是否含有‘top’,主要是去除出第一张图片
                    continue
                # print list
                if i < 10:  # 不足10 张用0补位
                    i = "0" + str(i)
                images_name = new_sku+"_" +str(i)+ ".jpg"
                i = int(i) + 1

                fpath = mkfile_time +'/'+ new_sku  # 指定放在D 盘 目录下
                # if not os.path.exists(su+'/'+new_sku):
                if not os.path.exists(fpath):
                    os.makedirs(fpath)  # makedirs 创建多级目录文件夹，mkdir创建一个文件夹
                r2 = requests.get(list)
                with open(fpath + '/' + images_name, "wb") as f:
                    f.write(r2.content)
                    print ('下载完成')

        #  下载产品主图

            image = d['images']
            i = 1
            for url in image:
                url_images = url.values()
                url_images = ''.join(url_images)
                images_name = new_sku+"_800" +str(i) + ".jpg"
                pack_name = new_sku+'-gallery'
                i = i + 1

                fpath = mkfile_time + '/' + pack_name  # 指定放在D 盘 目录下
                if not os.path.exists(fpath):  # 以 sku命名 的 大的 文件夹
                    os.makedirs(fpath)  # makedirs 创建多级目录文件夹，mkdir创建一个文件夹
                r2 = requests.get(url_images)
                with open(fpath + '/' + images_name, "wb") as f:
                    f.write(r2.content)
                    print ('下载完成')


    def down_zip(self, mypath):  # 下载文件打包
        zip_filename = mypath +'.zip'
        mypath = mypath + os.sep
        print mypath

        z = zipfile.ZipFile(zip_filename,'w', zipfile.ZIP_DEFLATED)
        for dirpath, dirnames, filenames in os.walk(mypath):
            fpath = dirpath.replace(mypath,'')
            fpath = fpath and  fpath + os.sep or ''
            for filename in filenames:
                z.write(os.path.join(dirpath, filename),fpath + filename)
            print ('压缩成功')
        z.close()


if __name__ == '__main__':
    # 按照规定格式输入sku列表，sku之间用逗号隔开，逗号必须为英文输入法模式下输入
    #serial = 'ZY0342000783,ZY0342000781'
    product_skus = raw_input("请输入产品编号列表:\n").split(',')
    object = Kebilin(product_skus)



